import sys
import subprocess
import csv
import os

# Try to install openpyxl with --break-system-packages if missing
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    has_openpyxl = True
except ImportError:
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        has_openpyxl = True
    except Exception as e:
        print(f"Could not install openpyxl: {e}. Will generate CSV files instead.")
        has_openpyxl = False

base_dir = "/Users/bossktt/Library/CloudStorage/GoogleDrive-bossktt@gmail.com/My Drive/AI in healthcare/AI-Project/AI-advice"

# 1. Clinician Form Data
cols_clinician = [
    "Respondent_ID", "Date", "Interviewer_Name", "Role_Position", "Department", "Hospital_Type",
    "Avg_Time_Per_Case_Min", "Explanation_Time_Adequacy", "Pain_Points_Quotes",
    "Practical_Need_Rating", "If_No_System_Impact",
    "Top_Pros_Liked", "Hallucination_Risk_Concern_1to5", "Malpractice_Liability_Concern_1to5",
    "Review_Time_Concern_1to5", "Privacy_PDPA_Concern_1to5", "Must_Have_Verification_Feature",
    "Willingness_To_Adopt", "Main_Adoption_Blockers", "Thai_English_CodeSwitching_Pref",
    "ClosedLoop_ReadReceipt_Need", "Actionable_Product_Requirements"
]

rows_clinician = []
for i in range(1, 11):
    rows_clinician.append([f"DOC-{i:02d}", "", "", "แพทย์ทั่วไป/แพทย์เฉพาะทาง", "OPD", "รพ.รัฐ", "", "ไม่เพียงพอ", "", "จำเป็นอย่างยิ่ง", "", "", "4", "5", "4", "3", "ปุ่มกดฟังเสียงย้อนหลัง Linked Audio Evidence", "ยินดีใช้ทุกเคส", "", "แปลเป็นภาษาไทย ป.5", "จำเป็นต้องมี", ""])

# 2. Patient / Caregiver Form Data
cols_patient = [
    "Respondent_ID", "Date", "Interviewer_Name", "Respondent_Group", "Hospital_Type", "Medical_Condition",
    "Current_Pain_Points", "Current_Experience_Quotes",
    "Practical_Need_Rating", "If_No_System_Impact",
    "Preferred_Channel", "Trust_In_AI_1to5", "Privacy_Audio_Recording_Concern",
    "Willingness_To_Use", "ClosedLoop_Confirmation_Willingness",
    "TextToSpeech_Voice_Need", "ColorCoded_Med_Matrix_Need", "Caregiver_Checklist_Need",
    "Actionable_Product_Requirements"
]

rows_patient = []
for i in range(1, 11):
    rows_patient.append([f"PT-{i:02d}", "", "", "ผู้ดูแล/ลูกหลาน", "รพ.รัฐ", "โรคความดัน/เบาหวาน", "จำคำสั่งหมอไม่ได้/สับสนเรื่องยา", "", "จำเป็นอย่างยิ่ง", "", "LINE OA", "4", "ไม่กังวล", "ยินดีใช้", "ยินดีกดยืนยันทุกครั้ง", "จำเป็น", "จำเป็นอย่างยิ่ง", "จำเป็นอย่างยิ่ง", ""])

# 3. Scenario Scores Data
cols_scenario = [
    "Respondent_ID", "Group", "Scenario1_OPD3Min_Score", "Scenario1_Comments",
    "Scenario2_ED_RedFlags_Score", "Scenario2_Comments",
    "Scenario3_IPD_MedMatrix_Score", "Scenario3_Comments",
    "Scenario4_ClosedLoop_Confirm_Score", "Scenario4_Comments",
    "Overall_Scenario_Average"
]

rows_scenario = []
for i in range(1, 11):
    rows_scenario.append([f"DOC-{i:02d}", "แพทย์", "5", "", "5", "", "5", "", "4", "", "=AVERAGE(C2,E2,G2,I2)"])
for i in range(1, 11):
    rows_scenario.append([f"PT-{i:02d}", "ผู้ดูแล/ผู้ป่วย", "5", "", "5", "", "5", "", "5", "", "=AVERAGE(C12,E12,G12,I12)"])

# Export to CSV with UTF-8 BOM for Excel / Google Sheets compatibility
def write_csv(filename, headers, rows):
    path = os.path.join(base_dir, filename)
    with open(path, mode='w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"Successfully created CSV: {path}")

write_csv("interview_form_clinician.csv", cols_clinician, rows_clinician)
write_csv("interview_form_patient_caregiver.csv", cols_patient, rows_patient)
write_csv("interview_form_scenarios.csv", cols_scenario, rows_scenario)

# Export to XLSX if openpyxl is ready
if has_openpyxl:
    wb = openpyxl.Workbook()
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_fill_green = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_fill_purple = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    
    regular_font = Font(name="Calibri", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color="CBD5E1"),
        right=Side(style='thin', color="CBD5E1"),
        top=Side(style='thin', color="CBD5E1"),
        bottom=Side(style='thin', color="CBD5E1")
    )
    
    # Sheet 1: Clinician
    ws1 = wb.active
    ws1.title = "Clinician_Form"
    ws1.append(cols_clinician)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = align_center
        cell.border = thin_border
    for r in rows_clinician:
        ws1.append(r)
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_center if cell.column in [1, 2, 5, 6, 7, 8, 10, 13, 14, 15, 16, 18] else align_left

    # Sheet 2: Patient/Caregiver
    ws2 = wb.create_sheet(title="Patient_Caregiver_Form")
    ws2.append(cols_patient)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill_green
        cell.alignment = align_center
        cell.border = thin_border
    for r in rows_patient:
        ws2.append(r)
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_center if cell.column in [1, 2, 4, 5, 9, 11, 12, 13, 14, 15, 16, 17, 18] else align_left

    # Sheet 3: Scenarios
    ws3 = wb.create_sheet(title="Scenario_Scores")
    ws3.append(cols_scenario)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill_purple
        cell.alignment = align_center
        cell.border = thin_border
    for r in rows_scenario:
        ws3.append(r)
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = align_center if cell.column in [1, 2, 3, 5, 7, 9, 11] else align_left

    # Column Auto-Width & Freeze Panes
    for sheet in wb.worksheets:
        sheet.freeze_panes = "B2"
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 18)

    excel_path = os.path.join(base_dir, "interview_form.xlsx")
    wb.save(excel_path)
    print(f"Successfully generated Excel file at: {excel_path}")
